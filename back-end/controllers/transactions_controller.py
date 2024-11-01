from flask import Blueprint, request, send_file
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from flask_cors import cross_origin
from connector.mysql_connector import connection
from models.users import UserModel
from models.cars import CarModel
from models.car_categories import CarCategoryModel
from models.drivers import DriverModel
from models.transactions import TransactionModel
from sqlalchemy.orm import sessionmaker
from flask_jwt_extended import jwt_required, get_jwt_identity
from cerberus import Validator
from schemas.transactions_schema import (
    add_transaction_schema,
    validating_payment_schema,
    return_car_schema,
    generate_report_schema,
)
from utils.handle_response import ResponseHandler
import os
import cloudinary
import cloudinary.uploader
from sqlalchemy import or_, and_, func
from datetime import datetime, timedelta

transactions_blueprint = Blueprint("transactions_blueprint", __name__)

cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET"),
)


@transactions_blueprint.post("/transactions")
@cross_origin(origin="localhost", headers=["Content-Type", "Authorization"])
@jwt_required()
def create_transaction():
    Session = sessionmaker(bind=connection)
    s = Session()
    s.begin()

    try:
        user_id = get_jwt_identity()
        current_user = s.query(UserModel).filter_by(id=user_id).first()
        if not current_user:
            return ResponseHandler.error(message="User not found", status=404)

        data = request.get_json()
        validator = Validator(add_transaction_schema)
        if not validator.validate(data):
            return ResponseHandler.error(message="Data Invalid!", data=validator.errors, status=400)

        car_name = data.get("car_name")
        driver_name = data.get("driver_name", None)  # Use None if driver is not selected
        start_date_string = data.get("start_date")
        end_date_string = data.get("end_date")

        # Convert the strings to datetime objects
        start_date = datetime.strptime(start_date_string, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_string, "%Y-%m-%d")

        if end_date <= start_date:
            return ResponseHandler.error(message="End date must be greater than start date", status=400)

        # Calculate rent duration
        date_difference = end_date - start_date
        rent_duration = date_difference.days
        driver_cost = 0

        existing_car = s.query(CarModel).filter(CarModel.name == car_name).first()
        if not existing_car:
            return ResponseHandler.error(message="Car not found!", status=404)
        if existing_car.status != "Available":
            return ResponseHandler.error(message="Car is not available!", status=404)

        # Check if the user want to use driver or not
        existing_driver = None
        if "driver_name" in data:
            existing_driver = s.query(DriverModel).filter(DriverModel.name == driver_name).first()
            if not existing_driver:
                return ResponseHandler.error(message="Driver not found!", status=404)
            if existing_driver.status != "Available":
                return ResponseHandler.error(message="Driver is not available!", status=404)

            # Calculate driver cost
            driver_cost = 100000 * rent_duration

        # Calculate total cost for transaction
        total_cost = (existing_car.price * rent_duration) + driver_cost

        # Create the transaction
        new_transaction = TransactionModel(
            user_id=user_id,
            car_id=existing_car.id,
            driver_id=existing_driver.id if existing_driver else None,
            start_date=start_date,
            end_date=end_date,
            return_date=None,
            rental_status="Pending",
            payment_status="Pending",
            payment_proof=None,
            late_fee=None,
            total_cost=total_cost,
        )
        new_transaction.generate_invoice()

        # Update car status to booked & Driver status if using driver
        existing_car.status = "Booked"
        if existing_driver:
            existing_driver.status = "Booked"

        s.add(new_transaction)
        s.commit()

        return ResponseHandler.success(
            message="Transaction added successfully",
            data=new_transaction.to_dictionaries(),
            status=201,
        )

    except Exception as e:
        s.rollback()
        return ResponseHandler.error(
            message="An error occurred while adding new transaction",
            data=str(e),
            status=500,
        )

    finally:
        s.close()


@transactions_blueprint.get("/transactions/customer")
@cross_origin(origin="localhost", headers=["Content-Type", "Authorization"])
@jwt_required()
# Show transactions belongs to current customer
def show_customer_transaction():
    user_id = get_jwt_identity()
    try:
        current_user = UserModel.query.filter_by(id=user_id).first()
        if not current_user:
            return ResponseHandler.error(message="User not found", status=404)

        # Get query parameters for pagination
        page = request.args.get("page", default=1, type=int)
        per_page = request.args.get("per_page", default=5, type=int)

        # Check the transactions belongs to the current user
        transactions = TransactionModel.query.filter_by(user_id=user_id).paginate(
            page=page, per_page=per_page, error_out=False
        )
        if not transactions:
            return ResponseHandler.error(message="No transactions found", status=404)

        # Prepare transactions data
        transactions_data = []
        for transaction in transactions:
            # Retrieve the car information for this transactiion
            car_info = CarModel.query.filter_by(id=transaction.car_id).first()
            category_info = CarCategoryModel.query.filter_by(id=car_info.category_id).first()
            car_data = {
                "car_slug": car_info.slug,
                "car_name": car_info.name,
                "car_price": car_info.price,
                "car_image": car_info.image,
                "car_brand": category_info.car_brand,
                "car_type": category_info.type,
            }

            # Retrieve the driver information for this transaction if there's any
            driver_data = None
            driver_info = DriverModel.query.filter_by(id=transaction.driver_id).first()
            if driver_info:
                driver_data = {"driver_name": driver_info.name, "driver_phone_number": driver_info.phone_number}

            # Append datas into transactions_data
            transactions_data.append(
                {
                    **{
                        column.name: getattr(transaction, column.name) for column in TransactionModel.__table__.columns
                    },
                    "car_data": car_data,
                    "driver_data": driver_data,
                }
            )

        response_data = {
            "transactions": transactions_data,
            "pagination": {
                "total_transactions": transactions.total,
                "current_page": transactions.page,
                "total_pages": transactions.pages,
                "next_page": page + 1 if page < transactions.pages else None,
                "prev_page": page - 1 if page > 1 else None,
            },
        }

        return ResponseHandler.success(data=response_data, status=200)

    except Exception as e:
        return ResponseHandler.error(
            message="An error occured while showing transactions",
            data=str(e),
            status=500,
        )


@transactions_blueprint.get("/transactions/admin")
@cross_origin(origin="localhost", headers=["Content-Type", "Authorization"])
@jwt_required()
# Only admin can show all transactions
def show_all_transaction():
    user_id = get_jwt_identity()
    try:
        current_user = UserModel.query.filter_by(id=user_id).first()
        if not current_user:
            return ResponseHandler.error(message="User not found", status=404)

        # Check if the current user's role is "admin"
        if current_user.role_id != 1:
            return ResponseHandler.error(message="Unauthorized access, only admin can access this!", status=403)

        # Get query parameters for pagination
        page = request.args.get("page", default=1, type=int)
        per_page = request.args.get("per_page", default=5, type=int)

        # Get query parameters for search filters
        rental_status = request.args.get("rental_status", default=None, type=str)
        payment_status = request.args.get("payment_status", default=None, type=str)

        query = TransactionModel.query

        # Apply search filters
        if rental_status:
            query = query.filter(TransactionModel.rental_status.ilike(f"%{rental_status}%"))

        if payment_status:
            query = query.filter(TransactionModel.payment_status.ilike(f"%{payment_status}%"))

        transactions = query.paginate(page=page, per_page=per_page, error_out=False)
        if not transactions:
            return ResponseHandler.error(message="No transactions found", status=404)

        # Prepare transactions data
        transactions_data = []
        for transaction in transactions:
            # Retrieve the car information for this transactiion
            car_info = CarModel.query.filter_by(id=transaction.car_id).first()
            category_info = CarCategoryModel.query.filter_by(id=car_info.category_id).first()
            car_data = {
                "car_slug": car_info.slug,
                "car_name": car_info.name,
                "car_price": car_info.price,
                "car_image": car_info.image,
                "car_brand": category_info.car_brand,
                "car_type": category_info.type,
            }

            # Retrieve the driver information for this transaction if there's any
            driver_data = None
            driver_info = DriverModel.query.filter_by(id=transaction.driver_id).first()
            if driver_info:
                driver_data = {"driver_name": driver_info.name, "driver_phone_number": driver_info.phone_number}

            # Append datas into transactions_data
            transactions_data.append(
                {
                    **{
                        column.name: getattr(transaction, column.name) for column in TransactionModel.__table__.columns
                    },
                    "car_data": car_data,
                    "driver_data": driver_data,
                }
            )

        response_data = {
            "transactions": transactions_data,
            "pagination": {
                "total_transactions": transactions.total,
                "current_page": transactions.page,
                "total_pages": transactions.pages,
                "next_page": page + 1 if page < transactions.pages else None,
                "prev_page": page - 1 if page > 1 else None,
            },
        }

        return ResponseHandler.success(data=response_data, status=200)

    except Exception as e:
        return ResponseHandler.error(
            message="An error occured while showing transactions",
            data=str(e),
            status=500,
        )


@transactions_blueprint.get("/transactions/<int:transaction_id>")
@cross_origin(origin="localhost", headers=["Content-Type", "Authorization"])
@jwt_required()
def show_transaction_by_id(transaction_id):
    user_id = get_jwt_identity()
    try:
        current_user = UserModel.query.filter_by(id=user_id).first()
        if not current_user:
            return ResponseHandler.error(message="User not found", status=404)

        transaction = TransactionModel.query.filter_by(id=transaction_id).first()
        if not transaction:
            return ResponseHandler.error(message="No transactions found", status=404)

        # Retrieve the car information for this transactiion
        car_info = CarModel.query.filter_by(id=transaction.car_id).first()
        category_info = CarCategoryModel.query.filter_by(id=car_info.category_id).first()
        car_data = {
            "car_slug": car_info.slug,
            "car_name": car_info.name,
            "car_price": car_info.price,
            "car_image": car_info.image,
            "car_brand": category_info.car_brand,
            "car_type": category_info.type,
        }

        # Retrieve the driver information for this transaction if there's any
        driver_data = None
        driver_info = DriverModel.query.filter_by(id=transaction.driver_id).first()
        if driver_info:
            driver_data = {"driver_name": driver_info.name, "driver_phone_number": driver_info.phone_number}

        transaction_data = {
            **{column.name: getattr(transaction, column.name) for column in TransactionModel.__table__.columns},
            "car_data": car_data,
            "driver_data": driver_data,
        }

        return ResponseHandler.success(data=transaction_data, status=200)

    except Exception as e:
        return ResponseHandler.error(
            message="An error occured while showing transaction",
            data=str(e),
            status=500,
        )


@transactions_blueprint.put("/transactions/upload-payment-proof/<int:transaction_id>")
@cross_origin(origin="localhost", headers=["Content-Type", "Authorization"])
@jwt_required()
# Upload payment proof for customer only
def upload_payment(transaction_id):
    Session = sessionmaker(bind=connection)
    s = Session()
    s.begin()

    try:
        user_id = get_jwt_identity()
        current_user = s.query(UserModel).filter_by(id=user_id).first()
        if not current_user:
            return ResponseHandler.error(message="User not found", status=404)

        # Check if the current user's role is "customer"
        if current_user.role_id != 2:
            return ResponseHandler.error(message="Unauthorized access, only customer can access this!", status=403)

        # Check transaction's data in database
        transaction = s.query(TransactionModel).filter_by(id=transaction_id, user_id=user_id).first()
        if not transaction:
            return ResponseHandler.error(message="Transaction not found or belongs to other users!", status=404)

        # Check if the payment proof is uploaded
        if transaction.payment_proof not in [None, ""]:
            return ResponseHandler.error(message="Payment Proof is uploaded!", status=400)

        # Get the uploaded files from the request
        files = request.files.getlist("payment_proof_image")
        if not files:
            return ResponseHandler.error(message="Payment Proof Image is required", status=400)

        image_urls = []

        for file in files:
            upload_result = cloudinary.uploader.upload(file)
            image_url = upload_result["secure_url"]
            image_urls.append(image_url)

        transaction.payment_proof = image_urls[0]
        s.commit()

        return ResponseHandler.success(
            message="Payment Proof successfully uploaded!",
            data=transaction.to_dictionaries(),
            status=200,
        )

    except Exception as e:
        s.rollback()
        return ResponseHandler.error(
            message="An error occurred while uploading the payment proof",
            data=str(e),
            status=500,
        )

    finally:
        s.close()


@transactions_blueprint.put("/transactions/payment-proof-validation/<int:transaction_id>")
@cross_origin(origin="localhost", headers=["Content-Type", "Authorization"])
@jwt_required()
# Only admin can checks whether the payment proof is valid or not
def payment_validation(transaction_id):
    Session = sessionmaker(bind=connection)
    s = Session()
    s.begin()

    try:
        user_id = get_jwt_identity()
        current_user = s.query(UserModel).filter_by(id=user_id).first()
        if not current_user:
            return ResponseHandler.error(message="User not found", status=404)

        # Check if the current user's role is "admin"
        if current_user.role_id != 1:
            return ResponseHandler.error(message="Unauthorized access, only admin can access this!", status=403)

        # Check transaction's data in database
        transaction = s.query(TransactionModel).filter_by(id=transaction_id).first()
        if not transaction:
            return ResponseHandler.error(message="Transaction not found!", status=404)

        data = request.get_json()
        validator = Validator(validating_payment_schema)
        if not validator.validate(data):
            return ResponseHandler.error(message="Invalid data!", data=validator.errors, status=400)

        rental_status = data.get("rental_status")

        if rental_status == "Valid":
            if transaction.driver_id not in [None, ""]:
                driver = s.query(DriverModel).filter_by(id=transaction.driver_id).first()
                driver.status = "Rented"

            car = s.query(CarModel).filter_by(id=transaction.car_id).first()
            car.status = "Rented"
            transaction.payment_status = "Success"
            transaction.rental_status = "In Progress"

        elif rental_status == "Invalid":
            if transaction.driver_id not in [None, ""]:
                driver = s.query(DriverModel).filter_by(id=transaction.driver_id).first()
                driver.status = "Available"

            car = s.query(CarModel).filter_by(id=transaction.car_id).first()
            car.status = "Available"
            transaction.payment_status = "Invalid"
            transaction.rental_status = "Canceled"

        s.commit()

        return ResponseHandler.success(
            message="Payment Proof validating is succeed!",
            data=transaction.to_dictionaries(),
            status=200,
        )

    except Exception as e:
        s.rollback()
        return ResponseHandler.error(
            message="An error occurred while validating the payment proof",
            data=str(e),
            status=500,
        )

    finally:
        s.close()


@transactions_blueprint.put("/transactions/return-car/<int:transaction_id>")
@cross_origin(origin="localhost", headers=["Content-Type", "Authorization"])
@jwt_required()
# Customer returning the car
def return_car(transaction_id):
    Session = sessionmaker(bind=connection)
    s = Session()
    s.begin()

    try:
        user_id = get_jwt_identity()
        current_user = s.query(UserModel).filter_by(id=user_id).first()
        if not current_user:
            return ResponseHandler.error(message="User not found", status=404)

        # Check if the current user's role is "customer"
        if current_user.role_id != 2:
            return ResponseHandler.error(message="Unauthorized access, only customer can access this!", status=403)

        # Check transaction's data in database
        transaction = s.query(TransactionModel).filter_by(id=transaction_id, user_id=user_id).first()
        if not transaction:
            return ResponseHandler.error(message="Transaction not found or belongs to other users!", status=404)

        if transaction.driver_id not in [None, ""]:
            driver = s.query(DriverModel).filter_by(id=transaction.driver_id).first()
            driver.status = "Available"

        car = s.query(CarModel).filter_by(id=transaction.car_id).first()
        car.status = "Available"

        data = request.get_json()
        validator = Validator(return_car_schema)
        if not validator.validate(data):
            return ResponseHandler.error(message="Invalid data!", data=validator.errors, status=400)

        return_date_string = data.get("return_date")

        return_date = datetime.strptime(return_date_string, "%Y-%m-%d").date()
        end_date = transaction.end_date

        if return_date < transaction.start_date:
            return ResponseHandler.error(message="Invalid return_date!", data=validator.errors, status=400)

        elif return_date > end_date:
            # Calculate days late & late_fee cost
            date_difference = return_date - end_date
            day_late = date_difference.days
            transaction.late_fee = 200000 * day_late
            transaction.total_cost += transaction.late_fee
            transaction.rental_status = "Success"

        elif return_date < end_date or return_date == end_date:
            transaction.late_fee = 0
            transaction.rental_status = "Success"

        transaction.return_date = return_date
        s.commit()

        return ResponseHandler.success(
            message="Return car success!",
            data=transaction.to_dictionaries(),
            status=200,
        )

    except Exception as e:
        s.rollback()
        return ResponseHandler.error(
            message="An error occurred while returning the car",
            data=str(e),
            status=500,
        )

    finally:
        s.close()


@transactions_blueprint.post("/transactions/generate_report")
@cross_origin(origin="localhost", headers=["Content-Type", "Authorization"])
@jwt_required()
def generate_report():
    Session = sessionmaker(bind=connection)
    s = Session()
    s.begin()

    def format_currency(amount):
        """Helper function to format number as IDR currency"""
        return f"Rp {'{:,.0f}'.format(amount).replace(',', '.')}"

    try:
        user_id = get_jwt_identity()
        current_user = s.query(UserModel).filter_by(id=user_id).first()
        if not current_user:
            return ResponseHandler.error(message="User not found", status=404)

        # Check if the current user's role is "admin"
        if current_user.role_id != 1:
            return ResponseHandler.error(message="Unauthorized access, only admin can access this!", status=403)

        data = request.get_json()
        validator = Validator(generate_report_schema)
        if not validator.validate(data):
            return ResponseHandler.error(message="Invalid data!", data=validator.errors, status=400)

        from_month = data.get("from_month")
        from_year = data.get("from_year")
        to_month = data.get("to_month")
        to_year = data.get("to_year")

        # Validate input months and convert them to month numbers
        try:
            start_date = datetime.strptime(f"01 {from_month} {from_year}", "%d %B %Y").date()
            end_date = datetime.strptime(f"01 {to_month} {to_year}", "%d %B %Y").date()

            # Adjust end_date to the last day of the month
            end_date = datetime(end_date.year, end_date.month + 1, 1) - timedelta(days=1)
            end_date = end_date.date()
        except ValueError as e:
            return ResponseHandler.error(
                message="Invalid month format!",
                data=str(e),
                status=400,
            )

        # Query transactions within the date range
        transactions = (
            s.query(TransactionModel)
            .filter(
                TransactionModel.end_date >= start_date,
                TransactionModel.end_date <= end_date,
                TransactionModel.rental_status == "Success",
            )
            .order_by(TransactionModel.end_date)
            .all()
        )

        if not transactions:
            return ResponseHandler.error(message="No transactions found for the given date range!", status=404)

        # Create an Excel workbook and get the active sheet
        wb = Workbook()
        wb.remove(wb.active)

        # Group transactions by month and year, using a sortable key
        monthly_transactions = {}
        for transaction in transactions:
            # Create a sortable key in the format YYYYMM
            sort_key = transaction.end_date.strftime("%Y%m")
            month_year = transaction.end_date.strftime("%B %Y")

            if sort_key not in monthly_transactions:
                monthly_transactions[sort_key] = {"display_name": month_year, "transactions": []}
            monthly_transactions[sort_key]["transactions"].append(transaction)

        # Create sheets for each month in chronological order
        for sort_key in sorted(monthly_transactions.keys()):
            month_data = monthly_transactions[sort_key]
            month_year = month_data["display_name"]
            month_transactions = month_data["transactions"]

            # Create a new sheet for each month
            ws = wb.create_sheet(title=f"{month_year} Transactions")

            # Set column widths
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 25
            ws.column_dimensions["D"].width = 25  # Increased width for currency column

            # Write headers with styling
            headers = ["Transaction ID", "User ID", "Invoice", "Total Cost"]
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

            # Write transaction data
            monthly_total = 0
            for transaction in month_transactions:
                ws.append(
                    [transaction.id, transaction.user_id, transaction.invoice, format_currency(transaction.total_cost)]
                )
                monthly_total += transaction.total_cost

            # Add empty row and monthly total
            ws.append([])
            ws.append([f"Total for {month_year}", "", "", format_currency(monthly_total)])

            # Style total row
            last_row = ws.max_row
            for cell in ws[last_row]:
                cell.font = Font(bold=True)

            # Add border to all cells
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

            # Align currency values to the right
            for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
                if row[3].value and isinstance(row[3].value, str) and row[3].value.startswith("Rp"):
                    row[3].alignment = Alignment(horizontal="right")

        # Create summary sheet
        summary_ws = wb.create_sheet(title="Summary", index=0)
        summary_ws.column_dimensions["A"].width = 25
        summary_ws.column_dimensions["B"].width = 25  # Increased width for currency column

        # Add summary headers
        summary_ws.append(["Month and Year", "Total Revenue"])
        for cell in summary_ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # Add monthly totals to summary in chronological order
        grand_total = 0
        for sort_key in sorted(monthly_transactions.keys()):
            month_data = monthly_transactions[sort_key]
            month_year = month_data["display_name"]
            month_transactions = month_data["transactions"]

            monthly_total = sum(t.total_cost for t in month_transactions)
            summary_ws.append([month_year, format_currency(monthly_total)])
            grand_total += monthly_total

        # Add empty row and grand total
        summary_ws.append([])
        summary_ws.append(["Grand Total", format_currency(grand_total)])
        for cell in summary_ws[summary_ws.max_row]:
            cell.font = Font(bold=True)

        # Add borders to summary sheet
        for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row):
            for cell in row:
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        # Align currency values to the right in summary sheet
        for row in summary_ws.iter_rows(min_row=2, max_col=2, max_row=summary_ws.max_row):
            if row[1].value and isinstance(row[1].value, str) and row[1].value.startswith("Rp"):
                row[1].alignment = Alignment(horizontal="right")

        # Save the Excel file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"transaction_report_{from_month}_to_{to_month}.xlsx"

        # Send the Excel file as a response
        response = send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        return response

    except Exception as e:
        s.rollback()
        return ResponseHandler.error(
            message="An error occurred while generating the transaction's report",
            data=str(e),
            status=500,
        )

    finally:
        s.close()
